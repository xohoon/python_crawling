from openpyxl import load_workbook

wb = load_workbook("C:/workspace/python/webcrawling/test.xlsx")
ws2 = wb.create_sheeet(title = "second_Sheet")
for row in range(1, 10):
    for col in range(1, 10):
        ws2.cell(row = row, column = col, value = int("{}{}".format(row, col)))

wb.save("C:/workspace/python/webcrawling/test.xlsx")